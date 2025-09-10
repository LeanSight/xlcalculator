#!/usr/bin/env python3
"""
Excel Range and Address Parsing Module

Centralized module for all Excel range and address parsing functionality.
Provides dataclasses and utilities for robust Excel reference handling.
"""

import collections
import re
from dataclasses import dataclass
from typing import Tuple, Optional, List, Union
from openpyxl.utils.cell import COORD_RE, SHEET_TITLE, range_boundaries, get_column_letter, column_index_from_string

# Import Excel constants
from .constants import EXCEL_MAX_COLUMNS, EXCEL_MAX_ROWS, EXCEL_MAX_COLUMN_INDEX

# Backward compatibility aliases
MAX_COL = EXCEL_MAX_COLUMN_INDEX
MAX_ROW = EXCEL_MAX_ROWS


def resolve_sheet(sheet_str: str) -> str:
    """Resolve sheet name from sheet string, handling quoted names."""
    sheet_str = sheet_str.strip()
    sheet_match = re.match(SHEET_TITLE.strip(), sheet_str + '!')
    if sheet_match is None:
        # Internally, sheets are not properly quoted, so consider the entire string
        return sheet_str
    return sheet_match.group("quoted") or sheet_match.group("notquoted")


@dataclass(frozen=True)
class CellReference:
    """Represents a cell reference with proper Excel context.
    
    Captures both explicit sheet references (Sheet1!A1) and 
    implicit references (A1) with their evaluation context.
    
    Attributes:
        sheet: Resolved sheet name
        address: Cell/range address (A1, A1:B2, etc.)
        is_sheet_explicit: Whether sheet was explicitly specified in original reference
    """
    sheet: str
    address: str
    is_sheet_explicit: bool
    
    @classmethod
    def parse(cls, ref: str, current_sheet: str = 'Sheet1') -> 'CellReference':
        """Parse reference string with proper sheet context.
        
        Args:
            ref: Reference string (e.g., 'Sheet1!A1' or 'A1')
            current_sheet: Current sheet context for implicit references
            
        Returns:
            CellReference object with resolved sheet and address
            
        Examples:
            CellReference.parse('Sheet1!A1', 'Sheet2') -> CellReference(sheet='Sheet1', address='A1', is_sheet_explicit=True)
            CellReference.parse('A1', 'Sheet2') -> CellReference(sheet='Sheet2', address='A1', is_sheet_explicit=False)
        """
        if '!' in ref:
            # Explicit sheet reference
            parts = ref.split('!', 1)
            sheet = resolve_sheet(parts[0])
            return cls(sheet=sheet, address=parts[1], is_sheet_explicit=True)
        else:
            # Implicit reference - use current sheet context
            return cls(sheet=current_sheet, address=ref, is_sheet_explicit=False)
    
    def __str__(self) -> str:
        """Return full sheet!address format."""
        return f"{self.sheet}!{self.address}"
    
    def is_same_sheet_as_context(self, context_sheet: str) -> bool:
        """Check if reference is in the same sheet as given context."""
        return self.sheet == context_sheet
    
    def to_tuple(self) -> Tuple[str, str]:
        """Return (sheet, address) tuple for backward compatibility."""
        return self.sheet, self.address


@dataclass(frozen=True)
class ParsedAddress:
    """Represents a parsed cell address with column and row components.
    
    Attributes:
        sheet: Sheet name
        column: Column letter(s) (A, B, AA, etc.)
        row: Row number
        full_address: Complete address string
    """
    sheet: str
    column: str
    row: int
    full_address: str
    
    @classmethod
    def parse(cls, addr: str) -> 'ParsedAddress':
        """Parse a complete address into components.
        
        Args:
            addr: Complete address like 'Sheet1!A1'
            
        Returns:
            ParsedAddress object with parsed components
            
        Raises:
            ValueError: If address format is invalid
        """
        if '!' not in addr:
            raise ValueError(f"Address must include sheet name: {addr}")
        
        sheet_str, addr_str = addr.split('!', 1)
        sheet = resolve_sheet(sheet_str)
        
        # Check for column reference (A:A)
        if ':' in addr_str and re.match(r'^[A-Z]+:[A-Z]+$', addr_str):
            # Column reference like A:A
            col_parts = addr_str.split(':')
            if col_parts[0] == col_parts[1]:
                # Single column reference
                return cls(sheet=sheet, column=col_parts[0], row=1, full_address=addr)
            else:
                raise ValueError(f"Multi-column ranges not supported yet: {addr_str}")
        
        # Check for row reference (1:1)
        if ':' in addr_str and re.match(r'^\d+:\d+$', addr_str):
            # Row reference like 1:1
            row_parts = addr_str.split(':')
            if row_parts[0] == row_parts[1]:
                # Single row reference
                return cls(sheet=sheet, column='A', row=int(row_parts[0]), full_address=addr)
            else:
                raise ValueError(f"Multi-row ranges not supported yet: {addr_str}")
        
        coord_match = COORD_RE.split(addr_str)
        if len(coord_match) < 3:
            raise ValueError(f"Invalid address format: {addr_str}")
        
        col, row = coord_match[1:3]
        return cls(sheet=sheet, column=col, row=int(row), full_address=addr)
    
    def __str__(self) -> str:
        """Return full address string."""
        return self.full_address


@dataclass(frozen=True)
class RangeReference:
    """Represents an Excel range reference with comprehensive parsing.
    
    Attributes:
        sheet: Sheet name
        address: Range address (A1:B2, A:A, 1:1, etc.)
        is_sheet_explicit: Whether sheet was explicitly specified
        is_full_column: Whether this is a full column reference (A:A)
        is_full_row: Whether this is a full row reference (1:1)
        min_col: Minimum column index (1-based)
        min_row: Minimum row index (1-based)
        max_col: Maximum column index (1-based)
        max_row: Maximum row index (1-based)
    """
    sheet: str
    address: str
    is_sheet_explicit: bool
    is_full_column: bool
    is_full_row: bool
    min_col: Optional[int]
    min_row: Optional[int]
    max_col: Optional[int]
    max_row: Optional[int]
    
    @classmethod
    def parse(cls, ref: str, current_sheet: str = 'Sheet1') -> 'RangeReference':
        """Parse range reference string with comprehensive analysis.
        
        Args:
            ref: Range reference string (e.g., 'Sheet1!A1:B2', 'A:A', '1:1')
            current_sheet: Current sheet context for implicit references
            
        Returns:
            RangeReference object with parsed components
        """
        # Parse sheet and address
        cell_ref = CellReference.parse(ref, current_sheet)
        sheet = cell_ref.sheet
        address = cell_ref.address
        is_sheet_explicit = cell_ref.is_sheet_explicit
        
        # Analyze range type and boundaries
        is_full_column = False
        is_full_row = False
        min_col = min_row = max_col = max_row = None
        
        if ':' in address:
            parts = address.split(':')
            if len(parts) == 2:
                left, right = parts
                
                # Check for full column (A:A, B:B)
                if left.isalpha() and right.isalpha() and left == right:
                    is_full_column = True
                    min_col = max_col = column_index_from_string(left)
                    min_row = 1
                    max_row = MAX_ROW
                
                # Check for full row (1:1, 2:2)
                elif left.isdigit() and right.isdigit() and left == right:
                    is_full_row = True
                    min_row = max_row = int(left)
                    min_col = 1
                    max_col = MAX_COL
                
                # Regular range (A1:B2)
                else:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(address)
                        # Handle unbound ranges
                        min_col = min_col or 1
                        min_row = min_row or 1
                        max_col = max_col or MAX_COL
                        max_row = max_row or MAX_ROW
                    except Exception:
                        # Fallback for invalid ranges
                        pass
        else:
            # Single cell reference
            try:
                coord_match = COORD_RE.split(address)
                if len(coord_match) >= 3:
                    col, row = coord_match[1:3]
                    min_col = max_col = column_index_from_string(col)
                    min_row = max_row = int(row)
            except Exception:
                pass
        
        return cls(
            sheet=sheet,
            address=address,
            is_sheet_explicit=is_sheet_explicit,
            is_full_column=is_full_column,
            is_full_row=is_full_row,
            min_col=min_col,
            min_row=min_row,
            max_col=max_col,
            max_row=max_row
        )
    
    def __str__(self) -> str:
        """Return full sheet!address format."""
        return f"{self.sheet}!{self.address}"
    
    def is_full_range(self) -> bool:
        """Check if this is a full column or row reference."""
        return self.is_full_column or self.is_full_row
    
    def to_cell_reference(self) -> CellReference:
        """Convert to CellReference for backward compatibility."""
        return CellReference(
            sheet=self.sheet,
            address=self.address,
            is_sheet_explicit=self.is_sheet_explicit
        )


# Backward compatibility functions
def parse_sheet_and_address(ref: str, default_sheet: str = 'Sheet1') -> Tuple[str, str]:
    """Parse reference into sheet name and address part.
    
    DEPRECATED: Use CellReference.parse() or RangeReference.parse() instead.
    
    Args:
        ref: Reference string (e.g., 'Sheet1!A1' or 'A1')
        default_sheet: Default sheet name when no sheet prefix provided
        
    Returns:
        Tuple of (sheet_name, address_part)
    """
    cell_ref = CellReference.parse(ref, current_sheet=default_sheet)
    return cell_ref.sheet, cell_ref.address


def resolve_address(addr: str) -> Tuple[str, str, str]:
    """Resolve address into sheet, column, and row components.
    
    DEPRECATED: Use ParsedAddress.parse() instead.
    
    Args:
        addr: Complete address like 'Sheet1!A1'
        
    Returns:
        Tuple of (sheet, column, row)
    """
    parsed = ParsedAddress.parse(addr)
    return parsed.sheet, parsed.column, str(parsed.row)


def resolve_ranges(ranges: str, default_sheet: str = 'Sheet1') -> Tuple[str, List[List[str]]]:
    """Resolve ranges string into sheet and cell matrix.
    
    Args:
        ranges: Comma-separated ranges string
        default_sheet: Default sheet name
        
    Returns:
        Tuple of (sheet_name, cell_matrix)
    """
    sheet = None
    range_cells = collections.defaultdict(set)
    
    for rng in ranges.split(','):
        # Parse range
        range_ref = RangeReference.parse(rng.strip(), default_sheet)
        
        # Validate sheet consistency
        if sheet is not None and sheet != range_ref.sheet:
            raise ValueError(
                f'Got multiple different sheets in ranges: {sheet}, {range_ref.sheet}'
            )
        sheet = range_ref.sheet
        
        # Add cells to matrix
        if range_ref.min_col and range_ref.min_row and range_ref.max_col and range_ref.max_row:
            for row_idx in range(range_ref.min_row, range_ref.max_row + 1):
                row_cells = range_cells[row_idx]
                for col_idx in range(range_ref.min_col, range_ref.max_col + 1):
                    row_cells.add(col_idx)
    
    # Convert to cell address matrix
    sheet = default_sheet if sheet is None else sheet
    sheet_str = sheet + '!' if sheet else ''
    
    return sheet, [
        [
            f'{sheet_str}{get_column_letter(col_idx)}{row_idx}'
            for col_idx in sorted(row_cells)
        ]
        for row_idx, row_cells in sorted(range_cells.items())
    ]


def is_full_range(range_str: str) -> bool:
    """Check if a range reference is a full column/row that needs lazy loading.
    
    Args:
        range_str: Range reference string
        
    Returns:
        True if this is a full column/row reference
    """
    try:
        range_ref = RangeReference.parse(range_str)
        return range_ref.is_full_range()
    except Exception:
        return False


# Legacy aliases for backward compatibility
def create_cell_reference(ref: str, current_sheet: str = 'Sheet1') -> CellReference:
    """Create CellReference object - alias for CellReference.parse()."""
    return CellReference.parse(ref, current_sheet)


def create_range_reference(ref: str, current_sheet: str = 'Sheet1') -> RangeReference:
    """Create RangeReference object - alias for RangeReference.parse()."""
    return RangeReference.parse(ref, current_sheet)