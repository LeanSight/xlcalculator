"""
Range Class Implementation

Represents an Excel range with proper cell collection,
address parsing, and range operations.
"""
from dataclasses import dataclass
from typing import TYPE_CHECKING, List, Optional, Tuple
from openpyxl.utils.cell import get_column_letter  # Removed column_index_from_string - using ParsedAddress data directly

if TYPE_CHECKING:
    from .worksheet import Worksheet
    from .cell import Cell


@dataclass
class Range:
    """Excel Range with hierarchical structure and cell management."""
    
    address: str  # Local range address like "A1:B2"
    worksheet: 'Worksheet'
    
    def __post_init__(self):
        """Initialize range after creation."""
        self._validate_address()
    
    def _validate_address(self) -> None:
        """Validate that the range address is in correct format."""
        if ':' not in self.address:
            raise ValueError(f"Invalid range address '{self.address}'. Expected format 'A1:B2'")
        
        try:
            start_addr, end_addr = self.address.split(':', 1)
            self._parse_cell_address(start_addr)
            self._parse_cell_address(end_addr)
        except Exception as e:
            raise ValueError(f"Invalid range address '{self.address}': {e}")
    
    def _parse_cell_address(self, address: str) -> Tuple[int, int]:
        """Parse cell address to get row and column indices.
        
        Args:
            address: Cell address like "A1"
            
        Returns:
            Tuple of (row, column_index)
        """
        from ..range import ParsedAddress
        full_address = f"{self.worksheet.name}!{address}"
        parsed = ParsedAddress.parse(full_address)
        return parsed.row, parsed.column_index
    
    @property
    def full_address(self) -> str:
        """Get the full address including sheet name."""
        return f"{self.worksheet.name}!{self.address}"
    
    @property
    def start_cell_address(self) -> str:
        """Get the start cell address of the range."""
        return self.address.split(':')[0]
    
    @property
    def end_cell_address(self) -> str:
        """Get the end cell address of the range."""
        return self.address.split(':')[1]
    
    @property
    def start_row(self) -> int:
        """Get the starting row number."""
        row, _ = self._parse_cell_address(self.start_cell_address)
        return row
    
    @property
    def end_row(self) -> int:
        """Get the ending row number."""
        row, _ = self._parse_cell_address(self.end_cell_address)
        return row
    
    @property
    def start_column_index(self) -> int:
        """Get the starting column index."""
        _, col_idx = self._parse_cell_address(self.start_cell_address)
        return col_idx
    
    @property
    def end_column_index(self) -> int:
        """Get the ending column index."""
        _, col_idx = self._parse_cell_address(self.end_cell_address)
        return col_idx
    
    @property
    def row_count(self) -> int:
        """Get the number of rows in the range."""
        return self.end_row - self.start_row + 1
    
    @property
    def column_count(self) -> int:
        """Get the number of columns in the range."""
        return self.end_column_index - self.start_column_index + 1
    
    @property
    def cell_count(self) -> int:
        """Get the total number of cells in the range."""
        return self.row_count * self.column_count
    
    @property
    def cells(self) -> List[List['Cell']]:
        """Get 2D array of cells in the range.
        
        Returns:
            List of rows, where each row is a list of cells
        """
        result = []
        
        for row_num in range(self.start_row, self.end_row + 1):
            row_cells = []
            for col_idx in range(self.start_column_index, self.end_column_index + 1):
                col_letter = get_column_letter(col_idx)
                cell_address = f"{col_letter}{row_num}"
                cell = self.worksheet.get_cell(cell_address)
                row_cells.append(cell)
            result.append(row_cells)
        
        return result
    
    def get_cell(self, row: int, col: int) -> 'Cell':
        """Get cell at specific position within the range.
        
        Args:
            row: Row index within range (0-based)
            col: Column index within range (0-based)
            
        Returns:
            Cell at the specified position
            
        Raises:
            IndexError: If position is outside range bounds
        """
        if row < 0 or row >= self.row_count:
            raise IndexError(f"Row index {row} is out of range (0-{self.row_count-1})")
        if col < 0 or col >= self.column_count:
            raise IndexError(f"Column index {col} is out of range (0-{self.column_count-1})")
        
        actual_row = self.start_row + row
        actual_col_idx = self.start_column_index + col
        actual_col_letter = get_column_letter(actual_col_idx)
        cell_address = f"{actual_col_letter}{actual_row}"
        
        return self.worksheet.get_cell(cell_address)
    
    def get_cell_by_address(self, address: str) -> 'Cell':
        """Get cell by its address within the range.
        
        Args:
            address: Cell address like "B2"
            
        Returns:
            Cell at the specified address
            
        Raises:
            ValueError: If address is not within the range
        """
        if not self.contains_address(address):
            raise ValueError(f"Address '{address}' is not within range '{self.address}'")
        
        return self.worksheet.get_cell(address)
    
    def contains_address(self, address: str) -> bool:
        """Check if the range contains the given cell address.
        
        Args:
            address: Cell address to check
            
        Returns:
            True if address is within the range
        """
        try:
            row, col_idx = self._parse_cell_address(address)
            return (
                self.start_row <= row <= self.end_row
                and self.start_column_index <= col_idx <= self.end_column_index
            )
        except Exception:
            return False
    
    def get_values(self) -> List[List]:
        """Get 2D array of cell values in the range.
        
        Returns:
            List of rows, where each row is a list of values
        """
        return [[cell.value for cell in row] for row in self.cells]
    
    def set_values(self, values: List[List]) -> None:
        """Set values for all cells in the range.
        
        Args:
            values: 2D array of values to set
            
        Raises:
            ValueError: If dimensions don't match
        """
        if len(values) != self.row_count:
            raise ValueError(f"Value array has {len(values)} rows, but range has {self.row_count}")
        
        for row_idx, row_values in enumerate(values):
            if len(row_values) != self.column_count:
                raise ValueError(f"Row {row_idx} has {len(row_values)} values, but range has {self.column_count} columns")
            
            for col_idx, value in enumerate(row_values):
                cell = self.get_cell(row_idx, col_idx)
                cell.set_value(value)
    
    def clear_values(self) -> None:
        """Clear all values in the range."""
        for row in self.cells:
            for cell in row:
                cell.clear()
    
    def get_row(self, row_index: int) -> List['Cell']:
        """Get all cells in a specific row of the range.
        
        Args:
            row_index: Row index within range (0-based)
            
        Returns:
            List of cells in the row
        """
        if row_index < 0 or row_index >= self.row_count:
            raise IndexError(f"Row index {row_index} is out of range")
        
        return self.cells[row_index]
    
    def get_column(self, col_index: int) -> List['Cell']:
        """Get all cells in a specific column of the range.
        
        Args:
            col_index: Column index within range (0-based)
            
        Returns:
            List of cells in the column
        """
        if col_index < 0 or col_index >= self.column_count:
            raise IndexError(f"Column index {col_index} is out of range")
        
        return [row[col_index] for row in self.cells]
    
    def resize(self, new_end_address: str) -> 'Range':
        """Create a new range with different end address.
        
        Args:
            new_end_address: New end cell address
            
        Returns:
            New Range object with updated dimensions
        """
        new_address = f"{self.start_cell_address}:{new_end_address}"
        return Range(address=new_address, worksheet=self.worksheet)
    
    def offset(self, row_offset: int, col_offset: int) -> 'Range':
        """Create a new range offset from this one.
        
        Args:
            row_offset: Number of rows to offset
            col_offset: Number of columns to offset
            
        Returns:
            New Range object at the offset position
        """
        # Calculate new start position
        new_start_row = self.start_row + row_offset
        new_start_col_idx = self.start_column_index + col_offset
        
        # Calculate new end position
        new_end_row = self.end_row + row_offset
        new_end_col_idx = self.end_column_index + col_offset
        
        # Validate bounds
        if new_start_row < 1 or new_start_col_idx < 1:
            raise ValueError("Offset results in invalid range position")
        
        # Create new addresses
        new_start_col = get_column_letter(new_start_col_idx)
        new_end_col = get_column_letter(new_end_col_idx)
        new_address = f"{new_start_col}{new_start_row}:{new_end_col}{new_end_row}"
        
        return Range(address=new_address, worksheet=self.worksheet)
    
    def intersect(self, other_range: 'Range') -> Optional['Range']:
        """Get intersection with another range.
        
        Args:
            other_range: Range to intersect with
            
        Returns:
            New Range representing intersection, or None if no intersection
        """
        if self.worksheet != other_range.worksheet:
            return None
        
        # Calculate intersection bounds
        start_row = max(self.start_row, other_range.start_row)
        end_row = min(self.end_row, other_range.end_row)
        start_col = max(self.start_column_index, other_range.start_column_index)
        end_col = min(self.end_column_index, other_range.end_column_index)
        
        # Check if intersection exists
        if start_row > end_row or start_col > end_col:
            return None
        
        # Create intersection range
        start_col_letter = get_column_letter(start_col)
        end_col_letter = get_column_letter(end_col)
        intersection_address = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
        
        return Range(address=intersection_address, worksheet=self.worksheet)
    
    def is_single_cell(self) -> bool:
        """Check if range represents a single cell.
        
        Returns:
            True if range is a single cell
        """
        return self.start_cell_address == self.end_cell_address
    
    def is_single_row(self) -> bool:
        """Check if range represents a single row.
        
        Returns:
            True if range spans only one row
        """
        return self.start_row == self.end_row
    
    def is_single_column(self) -> bool:
        """Check if range represents a single column.
        
        Returns:
            True if range spans only one column
        """
        return self.start_column_index == self.end_column_index
    
    def __eq__(self, other) -> bool:
        """Compare ranges for equality."""
        if not isinstance(other, Range):
            return False
        
        return (
            self.address == other.address
            and self.worksheet == other.worksheet
        )
    
    def __repr__(self) -> str:
        """String representation of range."""
        return f"Range('{self.full_address}', {self.row_count}x{self.column_count})"