"""
Worksheet Class Implementation

Represents an Excel worksheet with cell and range management,
local address handling, and formula support.
"""
from dataclasses import dataclass, field
from typing import Dict, TYPE_CHECKING, Any, Optional
import logging

from .cell import Cell
from .range import Range

if TYPE_CHECKING:
    from .workbook import Workbook


@dataclass
class Worksheet:
    """Excel Worksheet with hierarchical cell structure."""
    
    name: str
    workbook: 'Workbook'
    cells: Dict[str, Cell] = field(default_factory=dict)  # Local addresses like "A1"
    ranges: Dict[str, Range] = field(default_factory=dict)  # Local ranges like "A1:B2"
    visible: bool = True
    
    def get_cell(self, address: str) -> Cell:
        """Get cell by local address, creating if it doesn't exist.
        
        Args:
            address: Local cell address like "A1"
            
        Returns:
            The Cell instance
        """
        if address not in self.cells:
            # Create empty cell if it doesn't exist
            self.cells[address] = Cell(address=address, worksheet=self)
        
        return self.cells[address]
    
    def set_cell_value(self, address: str, value: Any) -> None:
        """Set cell value by local address.
        
        Args:
            address: Local cell address like "A1"
            value: Value to set in the cell
        """
        cell = self.get_cell(address)
        
        # Handle formula values
        if isinstance(value, str) and value.startswith('='):
            from ..xltypes import XLFormula
            formula = XLFormula(value, sheet_name=self.name)
            
            # Build AST automatically for immediate evaluation capability
            self._build_formula_ast(formula)
            
            cell.formula = formula
            cell.value = None  # Formula cells have no direct value
        else:
            cell.value = value
            cell.formula = None  # Clear any existing formula
    
    def get_cell_value(self, address: str) -> Any:
        """Get cell value by local address.
        
        Args:
            address: Local cell address like "A1"
            
        Returns:
            The cell value
        """
        cell = self.get_cell(address)
        return cell.value
    
    def _build_formula_ast(self, formula):
        """Build AST for a formula with current workbook context.
        
        Args:
            formula: XLFormula instance to build AST for
        """
        try:
            from .. import parser
            
            # Collect defined names from workbook for parser context
            defined_names = {}
            if hasattr(self, 'workbook') and self.workbook:
                for name, obj in self.workbook.defined_names.items():
                    if hasattr(obj, 'full_address'):
                        defined_names[name] = obj.full_address
                    elif hasattr(obj, 'address'):
                        defined_names[name] = f"{self.name}!{obj.address}"
            
            # Build the AST - avoid variable name conflict
            formula_text = formula.formula
            formula.ast = parser.FormulaParser().parse(formula_text, defined_names)
            
        except Exception as e:
            import logging
            logging.warning(f"Failed to parse formula '{formula.formula}' in {self.name}: {e}")
            # Leave AST as None - formula will still exist but won't be evaluable
    
    def build_all_formula_ast(self):
        """Build AST for all formulas in this worksheet."""
        for cell in self.cells.values():
            if cell.formula and not cell.formula.ast:
                self._build_formula_ast(cell.formula)
    
    def get_range(self, address: str) -> Range:
        """Get range by local address, creating if it doesn't exist.
        
        Args:
            address: Local range address like "A1:B2"
            
        Returns:
            The Range instance
        """
        if address not in self.ranges:
            # Create range if it doesn't exist
            self.ranges[address] = Range(address=address, worksheet=self)
        
        return self.ranges[address]
    
    def get_full_address(self, local_address: str) -> str:
        """Convert local address to full address.
        
        Args:
            local_address: Local address like "A1" or "A1:B2"
            
        Returns:
            Full address like "Sheet1!A1" or "Sheet1!A1:B2"
        """
        return f"{self.name}!{local_address}"
    
    def has_cell(self, address: str) -> bool:
        """Check if cell exists at given address.
        
        Args:
            address: Local cell address like "A1"
            
        Returns:
            True if cell exists and has been set, False otherwise
        """
        return address in self.cells and self.cells[address].value is not None
    
    def delete_cell(self, address: str) -> None:
        """Delete cell at given address.
        
        Args:
            address: Local cell address like "A1"
        """
        if address in self.cells:
            del self.cells[address]
    
    def clear_cell(self, address: str) -> None:
        """Clear cell value and formula but keep the cell object.
        
        Args:
            address: Local cell address like "A1"
        """
        if address in self.cells:
            cell = self.cells[address]
            cell.value = None
            cell.formula = None
    
    def get_used_range(self) -> Optional[Range]:
        """Get the range containing all used cells.
        
        Returns:
            Range object representing the used range, or None if no cells
        """
        if not self.cells:
            return None
        
        # Find min/max row and column
        min_row = min_col = float('inf')
        max_row = max_col = 0
        
        for address in self.cells:
            cell = self.cells[address]
            if cell.value is not None or cell.formula is not None:
                min_row = min(min_row, cell.row)
                max_row = max(max_row, cell.row)
                min_col = min(min_col, cell.column_index)
                max_col = max(max_col, cell.column_index)
        
        if min_row == float('inf'):
            return None
        
        # Convert column indices back to letters
        from openpyxl.utils.cell import get_column_letter
        min_col_letter = get_column_letter(min_col)
        max_col_letter = get_column_letter(max_col)
        
        range_address = f"{min_col_letter}{min_row}:{max_col_letter}{max_row}"
        return self.get_range(range_address)
    
    def copy_cell(self, from_address: str, to_address: str) -> None:
        """Copy cell from one address to another within the worksheet.
        
        Args:
            from_address: Source cell address
            to_address: Destination cell address
        """
        if from_address not in self.cells:
            return
        
        source_cell = self.cells[from_address]
        dest_cell = self.get_cell(to_address)
        
        dest_cell.value = source_cell.value
        if source_cell.formula:
            # TODO: Adjust formula references for new location
            dest_cell.formula = source_cell.formula
    
    def move_cell(self, from_address: str, to_address: str) -> None:
        """Move cell from one address to another within the worksheet.
        
        Args:
            from_address: Source cell address
            to_address: Destination cell address
        """
        self.copy_cell(from_address, to_address)
        self.delete_cell(from_address)
    
    def get_cells_with_formulas(self) -> Dict[str, Cell]:
        """Get all cells that contain formulas.
        
        Returns:
            Dictionary of address -> Cell for cells with formulas
        """
        return {
            address: cell
            for address, cell in self.cells.items()
            if cell.formula is not None
        }
    
    def get_cells_with_values(self) -> Dict[str, Cell]:
        """Get all cells that contain values (non-formula).
        
        Returns:
            Dictionary of address -> Cell for cells with values
        """
        return {
            address: cell
            for address, cell in self.cells.items()
            if cell.value is not None and cell.formula is None
        }
    
    def validate_address(self, address: str) -> bool:
        """Validate if address is a valid Excel cell address.
        
        Args:
            address: Cell address to validate
            
        Returns:
            True if valid, False otherwise
        """
        try:
            from ..range import ParsedAddress
            ParsedAddress.parse(f"{self.name}!{address}")
            return True
        except Exception:
            return False
    
    def get_cell_count(self) -> int:
        """Get total number of cells in the worksheet.
        
        Returns:
            Number of cells
        """
        return len(self.cells)
    
    def get_formula_count(self) -> int:
        """Get number of cells with formulas.
        
        Returns:
            Number of formula cells
        """
        return len(self.get_cells_with_formulas())
    
    def get_value_count(self) -> int:
        """Get number of cells with values (non-formula).
        
        Returns:
            Number of value cells
        """
        return len(self.get_cells_with_values())
    
    def __eq__(self, other) -> bool:
        """Compare worksheets for equality."""
        if not isinstance(other, Worksheet):
            return False
        
        return (
            self.name == other.name
            and self.cells == other.cells
            and self.ranges == other.ranges
            and self.visible == other.visible
        )
    
    def __repr__(self) -> str:
        """String representation of worksheet."""
        cell_count = len(self.cells)
        formula_count = self.get_formula_count()
        return f"Worksheet(name='{self.name}', cells={cell_count}, formulas={formula_count}, visible={self.visible})"