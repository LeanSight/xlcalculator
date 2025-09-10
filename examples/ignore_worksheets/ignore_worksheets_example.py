#!/usr/bin/env python3
"""
Ignore Worksheets Example

Demonstrates how to ignore specific worksheets when loading Excel files.
This is useful for large Excel files where you only need data from specific sheets.
"""

from xlcalculator import ModelCompiler
from xlcalculator import Evaluator
import openpyxl
import tempfile
import os


def create_sample_excel():
    """Create a sample Excel file with multiple worksheets"""
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    
    wb = openpyxl.Workbook()
    
    # Main calculation sheet
    ws1 = wb.active
    ws1.title = "Calculations"
    ws1['A1'] = "Revenue"
    ws1['B1'] = 1000000
    ws1['A2'] = "Costs"
    ws1['B2'] = 750000
    ws1['A3'] = "Profit"
    ws1['B3'] = "=B1-B2"
    
    # Configuration sheet (we want to ignore this)
    ws2 = wb.create_sheet("Config")
    ws2['A1'] = "Debug Mode"
    ws2['B1'] = True
    ws2['A2'] = "Log Level"
    ws2['B2'] = "DEBUG"
    
    # Data sheet
    ws3 = wb.create_sheet("Data")
    ws3['A1'] = "Product"
    ws3['B1'] = "Sales"
    ws3['A2'] = "Widget A"
    ws3['B2'] = 500
    ws3['A3'] = "Widget B"
    ws3['B3'] = 300
    ws3['A4'] = "Total"
    ws3['B4'] = "=SUM(B2:B3)"
    
    # Temporary calculations (we want to ignore this)
    ws4 = wb.create_sheet("TempCalcs")
    ws4['A1'] = "Temp value"
    ws4['B1'] = 999999
    
    wb.save(temp_file.name)
    wb.close()
    
    return temp_file.name


def example_load_all_sheets():
    """Example: Load all worksheets (default behavior)"""
    print("=== Loading ALL worksheets ===")
    
    excel_file = create_sample_excel()
    
    try:
        compiler = ModelCompiler()
        model = compiler.read_and_parse_archive(excel_file)
        evaluator = Evaluator(model)
        
        # Access data from all sheets
        profit = evaluator.evaluate('Calculations!B3')
        total_sales = evaluator.evaluate('Data!B4')
        debug_mode = evaluator.evaluate('Config!B1')
        temp_value = evaluator.evaluate('TempCalcs!B1')
        
        print(f"Profit: ${profit}")
        print(f"Total Sales: {total_sales}")
        print(f"Debug Mode: {debug_mode}")
        print(f"Temp Value: {temp_value}")
        
        # Get sheet names from cell addresses
        cell_addresses = list(model.cells.keys())
        sheet_names = set(addr.split('!')[0] for addr in cell_addresses if '!' in addr)
        print(f"Loaded sheets: {list(sheet_names)}")
        
    finally:
        os.unlink(excel_file)


def example_ignore_specific_sheets():
    """Example: Ignore specific worksheets by name"""
    print("\n=== Ignoring specific worksheets ===")
    
    excel_file = create_sample_excel()
    
    try:
        compiler = ModelCompiler()
        # Ignore configuration and temporary calculation sheets
        model = compiler.read_and_parse_archive(
            excel_file,
            ignore_sheets=['Config', 'TempCalcs']
        )
        evaluator = Evaluator(model)
        
        # Access data from included sheets
        profit = evaluator.evaluate('Calculations!B3')
        total_sales = evaluator.evaluate('Data!B4')
        
        print(f"Profit: ${profit}")
        print(f"Total Sales: {total_sales}")
        
        # Get sheet names from cell addresses
        cell_addresses = list(model.cells.keys())
        sheet_names = set(addr.split('!')[0] for addr in cell_addresses if '!' in addr)
        print(f"Loaded sheets: {list(sheet_names)}")
        
        # Try to access ignored sheet (will fail)
        try:
            debug_mode = evaluator.evaluate('Config!B1')
            print(f"Debug Mode: {debug_mode}")
        except Exception as e:
            print(f"✓ Config sheet properly ignored: {type(e).__name__}")
        
        try:
            temp_value = evaluator.evaluate('TempCalcs!B1')
            print(f"Temp Value: {temp_value}")
        except Exception as e:
            print(f"✓ TempCalcs sheet properly ignored: {type(e).__name__}")
            
    finally:
        os.unlink(excel_file)


def example_performance_comparison():
    """Example: Performance comparison with and without ignoring sheets"""
    print("\n=== Performance Comparison ===")
    
    excel_file = create_sample_excel()
    
    try:
        import time
        
        # Load all sheets
        start_time = time.time()
        compiler = ModelCompiler()
        model_all = compiler.read_and_parse_archive(excel_file)
        time_all = time.time() - start_time
        
        # Load only needed sheets
        start_time = time.time()
        compiler = ModelCompiler()
        model_filtered = compiler.read_and_parse_archive(
            excel_file,
            ignore_sheets=['Config', 'TempCalcs']
        )
        time_filtered = time.time() - start_time
        
        # Get sheet counts
        all_addresses = list(model_all.cells.keys())
        all_sheet_names = set(addr.split('!')[0] for addr in all_addresses if '!' in addr)
        
        filtered_addresses = list(model_filtered.cells.keys())
        filtered_sheet_names = set(addr.split('!')[0] for addr in filtered_addresses if '!' in addr)
        
        print(f"Load all sheets: {time_all:.4f}s ({len(all_sheet_names)} sheets)")
        print(f"Load filtered: {time_filtered:.4f}s ({len(filtered_sheet_names)} sheets)")
        print(f"Performance improvement: {((time_all - time_filtered) / time_all * 100):.1f}%")
        
    finally:
        os.unlink(excel_file)


if __name__ == '__main__':
    print("xlcalculator - Ignore Worksheets Example")
    print("=" * 50)
    
    example_load_all_sheets()
    example_ignore_specific_sheets()
    example_performance_comparison()
    
    print("\n✅ Example completed successfully!")