#!/usr/bin/env python3
"""
ATDD Test for Ignore Worksheets functionality

Test-driven example demonstrating how to ignore specific worksheets
when loading Excel files with xlcalculator.
"""

import pytest
import tempfile
import os
from xlcalculator import ModelCompiler
from xlcalculator import Evaluator
import openpyxl


class TestIgnoreWorksheets:
    """Test cases for worksheet ignoring functionality"""
    
    def setup_method(self):
        """Create a test Excel file with multiple worksheets"""
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()
        
        # Create Excel file with multiple sheets
        wb = openpyxl.Workbook()
        
        # Sheet1 (default) - should be included
        ws1 = wb.active
        ws1.title = "MainData"
        ws1['A1'] = 100
        ws1['B1'] = 200
        ws1['C1'] = "=A1+B1"
        
        # Sheet2 - should be ignored
        ws2 = wb.create_sheet("IgnoreMe")
        ws2['A1'] = 999
        ws2['B1'] = "=A1*2"
        
        # Sheet3 - should be included
        ws3 = wb.create_sheet("SecondaryData")
        ws3['A1'] = 50
        ws3['B1'] = "=A1*3"
        
        # Sheet4 - should be ignored
        ws4 = wb.create_sheet("AlsoIgnore")
        ws4['A1'] = 777
        
        wb.save(self.temp_file.name)
        wb.close()
    
    def teardown_method(self):
        """Clean up test file"""
        if os.path.exists(self.temp_file.name):
            os.unlink(self.temp_file.name)
    
    def test_load_all_worksheets(self):
        """Test loading all worksheets (baseline)"""
        compiler = ModelCompiler()
        model = compiler.read_and_parse_archive(self.temp_file.name)
        evaluator = Evaluator(model)
        
        # Should be able to access all sheets
        assert evaluator.evaluate('MainData!A1') == 100
        assert evaluator.evaluate('MainData!C1') == 300
        assert evaluator.evaluate('IgnoreMe!A1') == 999
        assert evaluator.evaluate('IgnoreMe!B1') == 1998
        assert evaluator.evaluate('SecondaryData!A1') == 50
        assert evaluator.evaluate('SecondaryData!B1') == 150
        assert evaluator.evaluate('AlsoIgnore!A1') == 777
    
    def test_ignore_specific_worksheets(self):
        """Test ignoring specific worksheets by name"""
        compiler = ModelCompiler()
        model = compiler.read_and_parse_archive(
            self.temp_file.name,
            ignore_sheets=['IgnoreMe', 'AlsoIgnore']
        )
        evaluator = Evaluator(model)
        
        # Should be able to access included sheets
        assert evaluator.evaluate('MainData!A1') == 100
        assert evaluator.evaluate('MainData!C1') == 300
        assert evaluator.evaluate('SecondaryData!A1') == 50
        assert evaluator.evaluate('SecondaryData!B1') == 150
        
        # Check that ignored sheets are not in the model by checking cell addresses
        cell_addresses = list(model.cells.keys())
        sheet_names = set(addr.split('!')[0] for addr in cell_addresses if '!' in addr)
        
        assert 'IgnoreMe' not in sheet_names
        assert 'AlsoIgnore' not in sheet_names
        assert 'MainData' in sheet_names
        assert 'SecondaryData' in sheet_names
    
    def test_ignore_worksheets_pattern(self):
        """Test ignoring worksheets by pattern (manual pattern matching)"""
        # Since pattern matching might not be built-in, we'll test manual selection
        compiler = ModelCompiler()
        
        # Get all sheet names first
        temp_model = compiler.read_and_parse_archive(self.temp_file.name)
        temp_addresses = list(temp_model.cells.keys())
        all_sheets = set(addr.split('!')[0] for addr in temp_addresses if '!' in addr)
        
        # Manually filter sheets containing "Ignore"
        ignore_sheets = [sheet for sheet in all_sheets if 'Ignore' in sheet]
        
        model = compiler.read_and_parse_archive(
            self.temp_file.name,
            ignore_sheets=ignore_sheets
        )
        evaluator = Evaluator(model)
        
        # Should be able to access non-matching sheets
        assert evaluator.evaluate('MainData!A1') == 100
        assert evaluator.evaluate('SecondaryData!A1') == 50
        
        # Check that pattern-matched sheets are not in the model
        final_addresses = list(model.cells.keys())
        final_sheet_names = set(addr.split('!')[0] for addr in final_addresses if '!' in addr)
        
        assert 'IgnoreMe' not in final_sheet_names
        assert 'AlsoIgnore' not in final_sheet_names


if __name__ == '__main__':
    # Run the tests
    test_instance = TestIgnoreWorksheets()
    test_instance.setup_method()
    
    try:
        print("Testing ignore worksheets functionality...")
        
        print("✓ Test 1: Load all worksheets")
        test_instance.test_load_all_worksheets()
        
        print("✓ Test 2: Ignore specific worksheets")
        test_instance.test_ignore_specific_worksheets()
        
        print("✓ Test 3: Ignore worksheets by pattern")
        test_instance.test_ignore_worksheets_pattern()
        
        print("All tests passed! ✅")
        
    except Exception as e:
        print(f"Test failed: {e} ❌")
        
    finally:
        test_instance.teardown_method()