#!/usr/bin/env python3
"""
ATDD Test for Model Focusing

Tests only the functionality that actually works in xlcalculator:
- ignore_sheets parameter in read_and_parse_archive
"""

import pytest
import tempfile
import os
from xlcalculator import ModelCompiler
from xlcalculator import Evaluator
import openpyxl


class TestModelFocusing:
    """Test cases for model focusing using only working functionality"""

    def create_test_excel(self):
        """Create a test Excel file with multiple sheets"""
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        wb = openpyxl.Workbook()
        
        # Core calculations sheet
        ws_core = wb.active
        ws_core.title = "Core"
        ws_core['A1'] = "Base Value"
        ws_core['B1'] = 1000
        ws_core['A2'] = "Multiplier"
        ws_core['B2'] = 2.5
        ws_core['A3'] = "Result"
        ws_core['B3'] = "=B1*B2"
        
        # Large unnecessary sheet
        ws_large = wb.create_sheet("LargeData")
        for i in range(1, 101):
            ws_large[f'A{i}'] = f"Item {i}"
            ws_large[f'B{i}'] = i * 100
            ws_large[f'C{i}'] = f"=B{i}*2"
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary['A1'] = "Final Value"
        ws_summary['B1'] = "=Core!B3"
        ws_summary['A2'] = "Status"
        ws_summary['B2'] = "Complete"
        
        wb.save(temp_file.name)
        wb.close()
        
        return temp_file.name

    def test_ignore_sheets_parameter_works(self):
        """
        GIVEN an Excel file with multiple sheets including large unnecessary data
        WHEN I load the model using ignore_sheets parameter
        THEN the specified sheets should be excluded from the model
        AND calculations should work correctly for included sheets
        """
        excel_file = self.create_test_excel()
        
        try:
            # Load model with ignore_sheets parameter
            compiler = ModelCompiler()
            model = compiler.read_and_parse_archive(
                excel_file,
                ignore_sheets=['LargeData']
            )
            
            # Verify ignored sheet is not in the model
            cell_addresses = list(model.cells.keys())
            sheet_names = set(addr.split('!')[0] for addr in cell_addresses if '!' in addr)
            
            assert 'LargeData' not in sheet_names, "LargeData sheet should be ignored"
            assert 'Core' in sheet_names, "Core sheet should be included"
            assert 'Summary' in sheet_names, "Summary sheet should be included"
            
            # Verify calculations work for included sheets
            evaluator = Evaluator(model)
            core_result = evaluator.evaluate('Core!B3')
            final_value = evaluator.evaluate('Summary!B1')
            
            assert core_result == 2500.0, f"Expected 2500.0, got {core_result}"
            assert final_value == 2500.0, f"Expected 2500.0, got {final_value}"
            
            # Verify ignored sheet cells are not accessible
            large_data_cells = [addr for addr in cell_addresses if addr.startswith('LargeData!')]
            assert len(large_data_cells) == 0, "No LargeData cells should be in the model"
                
        finally:
            os.unlink(excel_file)

    def test_ignore_multiple_sheets(self):
        """
        GIVEN an Excel file with multiple sheets
        WHEN I ignore multiple sheets using ignore_sheets parameter
        THEN all specified sheets should be excluded
        """
        excel_file = self.create_test_excel()
        
        try:
            compiler = ModelCompiler()
            model = compiler.read_and_parse_archive(
                excel_file,
                ignore_sheets=['LargeData', 'Summary']
            )
            
            # Verify only Core sheet is loaded
            cell_addresses = list(model.cells.keys())
            sheet_names = set(addr.split('!')[0] for addr in cell_addresses if '!' in addr)
            
            assert 'Core' in sheet_names, "Core sheet should be included"
            assert 'LargeData' not in sheet_names, "LargeData sheet should be ignored"
            assert 'Summary' not in sheet_names, "Summary sheet should be ignored"
            
            # Verify Core calculations still work
            evaluator = Evaluator(model)
            core_result = evaluator.evaluate('Core!B3')
            assert core_result == 2500.0
            
        finally:
            os.unlink(excel_file)

    def test_model_size_reduction_with_ignore_sheets(self):
        """
        GIVEN a large Excel file
        WHEN I compare model size with and without ignore_sheets
        THEN the focused model should have significantly fewer cells
        """
        excel_file = self.create_test_excel()
        
        try:
            # Load full model
            compiler_full = ModelCompiler()
            model_full = compiler_full.read_and_parse_archive(excel_file)
            
            # Load focused model (ignoring large sheet)
            compiler_focused = ModelCompiler()
            model_focused = compiler_focused.read_and_parse_archive(
                excel_file,
                ignore_sheets=['LargeData']
            )
            
            # Verify significant size reduction
            full_cells = len(model_full.cells)
            focused_cells = len(model_focused.cells)
            
            assert focused_cells < full_cells, "Focused model should have fewer cells"
            
            # Calculate reduction percentage
            reduction_percentage = ((full_cells - focused_cells) / full_cells) * 100
            assert reduction_percentage > 50, f"Should have >50% reduction, got {reduction_percentage:.1f}%"
            
        finally:
            os.unlink(excel_file)

    def test_all_real_parameters_work(self):
        """
        GIVEN the ModelCompiler.read_and_parse_archive method
        WHEN I use all real parameters
        THEN they should work without errors
        """
        excel_file = self.create_test_excel()
        
        try:
            compiler = ModelCompiler()
            
            # Test all REAL parameters work
            model = compiler.read_and_parse_archive(
                file_name=excel_file,
                ignore_sheets=['LargeData'],
                ignore_hidden=False,
                build_code=True
            )
            
            # Verify model was created successfully
            assert model is not None, "Model should be created"
            assert len(model.cells) > 0, "Model should have cells"
            
            # Verify LargeData was ignored
            cell_addresses = list(model.cells.keys())
            sheet_names = set(addr.split('!')[0] for addr in cell_addresses if '!' in addr)
            assert 'LargeData' not in sheet_names, "LargeData should be ignored"
            
        finally:
            os.unlink(excel_file)