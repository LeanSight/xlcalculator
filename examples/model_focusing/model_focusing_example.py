#!/usr/bin/env python3
"""
Model Focusing Example

Demonstrates ACTUAL model focusing capabilities that work in xlcalculator:
1. ignore_sheets parameter - exclude specific sheets during loading
2. ignore_hidden parameter - exclude hidden sheets during loading  

This example uses ONLY verified and working functionality.
"""

from xlcalculator import ModelCompiler
from xlcalculator import Evaluator
import openpyxl
import tempfile
import os
import time


def create_financial_model():
    """Create a financial model with multiple sheets for demonstration"""
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    
    wb = openpyxl.Workbook()
    
    # Assumptions Sheet
    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"
    ws_assumptions['A1'] = "Financial Model Assumptions"
    ws_assumptions['A3'] = "Revenue Growth Rate"
    ws_assumptions['B3'] = 0.12  # 12%
    ws_assumptions['A4'] = "Cost of Goods Sold %"
    ws_assumptions['B4'] = 0.60  # 60%
    ws_assumptions['A5'] = "Operating Expense %"
    ws_assumptions['B5'] = 0.25  # 25%
    ws_assumptions['A6'] = "Tax Rate"
    ws_assumptions['B6'] = 0.21  # 21%
    ws_assumptions['A8'] = "Base Year Revenue"
    ws_assumptions['B8'] = 10000000  # $10M
    
    # Revenue Projections Sheet
    ws_revenue = wb.create_sheet("Revenue")
    ws_revenue['A1'] = "Revenue Projections"
    ws_revenue['A2'] = "Year"
    ws_revenue['B2'] = "Revenue"
    
    # 5-year projections
    for year in range(1, 6):
        ws_revenue[f'A{year+2}'] = f"Year {year}"
        if year == 1:
            ws_revenue[f'B{year+2}'] = "=Assumptions!B8"
        else:
            ws_revenue[f'B{year+2}'] = f"=B{year+1}*(1+Assumptions!B3)"
    
    # P&L Sheet
    ws_pl = wb.create_sheet("ProfitLoss")
    ws_pl['A1'] = "Profit & Loss Statement"
    ws_pl['A2'] = "Year"
    ws_pl['B2'] = "Revenue"
    ws_pl['C2'] = "COGS"
    ws_pl['D2'] = "OpEx"
    ws_pl['E2'] = "EBITDA"
    ws_pl['F2'] = "Tax"
    ws_pl['G2'] = "Net Income"
    
    for year in range(1, 6):
        ws_pl[f'A{year+2}'] = f"Year {year}"
        ws_pl[f'B{year+2}'] = f"=Revenue!B{year+2}"
        ws_pl[f'C{year+2}'] = f"=B{year+2}*Assumptions!B4"
        ws_pl[f'D{year+2}'] = f"=B{year+2}*Assumptions!B5"
        ws_pl[f'E{year+2}'] = f"=B{year+2}-C{year+2}-D{year+2}"
        ws_pl[f'F{year+2}'] = f"=E{year+2}*Assumptions!B6"
        ws_pl[f'G{year+2}'] = f"=E{year+2}-F{year+2}"
    
    # Large Monthly Data (sheet we want to exclude)
    ws_monthly = wb.create_sheet("MonthlyDetails")
    ws_monthly['A1'] = "Monthly Revenue Details"
    ws_monthly['A2'] = "Month"
    ws_monthly['B2'] = "Revenue"
    ws_monthly['C2'] = "Customers"
    
    # 60 months of data (5 years)
    for month in range(1, 61):
        ws_monthly[f'A{month+2}'] = f"Month {month}"
        ws_monthly[f'B{month+2}'] = 800000 + month * 5000
        ws_monthly[f'C{month+2}'] = 1000 + month * 10
    
    # Summary Dashboard
    ws_summary = wb.create_sheet("Dashboard")
    ws_summary['A1'] = "Executive Dashboard"
    ws_summary['A3'] = "Key Metrics"
    ws_summary['A4'] = "Year 5 Revenue"
    ws_summary['B4'] = "=Revenue!B7"
    ws_summary['A5'] = "Year 5 Net Income"
    ws_summary['B5'] = "=ProfitLoss!G7"
    ws_summary['A6'] = "5-Year Revenue CAGR"
    ws_summary['B6'] = "=POWER(Revenue!B7/Revenue!B3,1/4)-1"
    
    # Configuration sheet (we might want to ignore)
    ws_config = wb.create_sheet("Config")
    ws_config['A1'] = "Model Configuration"
    ws_config['A2'] = "Debug Mode"
    ws_config['B2'] = True
    ws_config['A3'] = "Calculation Mode"
    ws_config['B3'] = "Automatic"
    
    wb.save(temp_file.name)
    wb.close()
    
    return temp_file.name


def example_full_model_analysis():
    """Example: Analyze the full model"""
    print("=== Full Model Analysis ===")
    
    excel_file = create_financial_model()
    
    try:
        start_time = time.time()
        compiler = ModelCompiler()
        model = compiler.read_and_parse_archive(excel_file)
        evaluator = Evaluator(model)
        load_time = time.time() - start_time
        
        print(f"Model loaded in {load_time:.4f} seconds")
        print(f"Total cells in model: {len(model.cells)}")
        print(f"Total formulas: {len(model.formulae)}")
        
        # Get sheet breakdown
        sheet_counts = {}
        for cell_addr in model.cells.keys():
            if '!' in cell_addr:
                sheet = cell_addr.split('!')[0]
                sheet_counts[sheet] = sheet_counts.get(sheet, 0) + 1
        
        print("\nCells per sheet:")
        for sheet, count in sheet_counts.items():
            print(f"  {sheet}: {count} cells")
        
        # Evaluate key metrics
        print("\nKey Financial Metrics:")
        year5_revenue = evaluator.evaluate('Dashboard!B4')
        year5_net = evaluator.evaluate('Dashboard!B5')
        
        print(f"Year 5 Revenue: ${float(year5_revenue):,.0f}")
        print(f"Year 5 Net Income: ${float(year5_net):,.0f}")
        
    finally:
        os.unlink(excel_file)


def example_focused_model_with_ignore_sheets():
    """Example: Focus model by ignoring specific sheets"""
    print("\n=== Focused Model with ignore_sheets ===")
    
    excel_file = create_financial_model()
    
    try:
        compiler = ModelCompiler()
        
        # Load full model
        full_model = compiler.read_and_parse_archive(excel_file)
        
        # Load focused model (ignoring large and config sheets)
        focused_model = compiler.read_and_parse_archive(
            excel_file, 
            ignore_sheets=['MonthlyDetails', 'Config']
        )
        
        print(f"Full model cells: {len(full_model.cells)}")
        print(f"Focused model cells: {len(focused_model.cells)}")
        reduction = (1 - len(focused_model.cells) / len(full_model.cells)) * 100
        print(f"Model size reduction: {reduction:.1f}%")
        
        # Verify key calculations still work
        evaluator = Evaluator(focused_model)
        print("\nKey Metrics (Focused Model):")
        year5_revenue = evaluator.evaluate('Dashboard!B4')
        year5_net = evaluator.evaluate('Dashboard!B5')
        
        print(f"Year 5 Revenue: ${float(year5_revenue):,.0f}")
        print(f"Year 5 Net Income: ${float(year5_net):,.0f}")
        
        # Verify ignored sheets are not accessible
        cell_addresses = list(focused_model.cells.keys())
        sheet_names = set(addr.split('!')[0] for addr in cell_addresses if '!' in addr)
        
        print(f"\nLoaded sheets: {sorted(sheet_names)}")
        print(f"MonthlyDetails ignored: {'✓' if 'MonthlyDetails' not in sheet_names else '✗'}")
        print(f"Config ignored: {'✓' if 'Config' not in sheet_names else '✗'}")
        
    finally:
        os.unlink(excel_file)


def example_scenario_analysis_with_focusing():
    """Example: Scenario analysis with focused model"""
    print("\n=== Scenario Analysis with Focused Model ===")
    
    excel_file = create_financial_model()
    
    try:
        compiler = ModelCompiler()
        # Focus on core calculation sheets only
        model = compiler.read_and_parse_archive(
            excel_file, 
            ignore_sheets=['MonthlyDetails', 'Config']
        )
        evaluator = Evaluator(model)
        
        print("Scenario Analysis: Impact of Growth Rate Changes")
        print(f"{'Growth Rate':<12} {'Year 5 Revenue':<15} {'Year 5 Net Income':<18}")
        print("-" * 50)
        
        # Test different growth rate scenarios
        growth_rates = [0.08, 0.10, 0.12, 0.15, 0.18]  # 8% to 18%
        
        for rate in growth_rates:
            # Change the growth rate assumption
            evaluator.set_cell_value('Assumptions!B3', rate)
            
            # Recalculate key metrics
            year5_revenue = evaluator.evaluate('Dashboard!B4')
            year5_net = evaluator.evaluate('Dashboard!B5')
            
            print(f"{rate*100:>8.0f}%     ${float(year5_revenue):>12,.0f}   ${float(year5_net):>15,.0f}")
        
        # Reset to original value
        evaluator.set_cell_value('Assumptions!B3', 0.12)
        
    finally:
        os.unlink(excel_file)


def example_performance_comparison():
    """Example: Compare performance of full vs focused models"""
    print("\n=== Performance Comparison ===")
    
    excel_file = create_financial_model()
    
    try:
        compiler = ModelCompiler()
        
        # Test full model performance
        start_time = time.time()
        full_model = compiler.read_and_parse_archive(excel_file)
        full_evaluator = Evaluator(full_model)
        full_result = full_evaluator.evaluate('Dashboard!B4')
        full_time = time.time() - start_time
        
        # Test focused model performance
        start_time = time.time()
        focused_model = compiler.read_and_parse_archive(
            excel_file, 
            ignore_sheets=['MonthlyDetails', 'Config']
        )
        focused_evaluator = Evaluator(focused_model)
        focused_result = focused_evaluator.evaluate('Dashboard!B4')
        focused_time = time.time() - start_time
        
        print("Performance Comparison:")
        print(f"Full Model:")
        print(f"  Cells: {len(full_model.cells)}")
        print(f"  Load + Evaluate Time: {full_time:.4f}s")
        print(f"  Result: ${float(full_result):,.0f}")
        
        print(f"\nFocused Model:")
        print(f"  Cells: {len(focused_model.cells)}")
        print(f"  Load + Evaluate Time: {focused_time:.4f}s")
        print(f"  Result: ${float(focused_result):,.0f}")
        
        speedup = full_time / focused_time if focused_time > 0 else float('inf')
        size_reduction = (1 - len(focused_model.cells) / len(full_model.cells)) * 100
        
        print(f"\nImprovement:")
        print(f"  Size Reduction: {size_reduction:.1f}%")
        print(f"  Speed Improvement: {speedup:.1f}x faster")
        print(f"  Results Match: {'✓' if abs(float(full_result) - float(focused_result)) < 0.01 else '✗'}")
        
    finally:
        os.unlink(excel_file)


def example_ignore_hidden_sheets():
    """Example: Ignore hidden sheets"""
    print("\n=== Ignore Hidden Sheets ===")
    
    excel_file = create_financial_model()
    
    try:
        # First, hide the Config sheet
        wb = openpyxl.load_workbook(excel_file)
        wb['Config'].sheet_state = 'hidden'
        wb.save(excel_file)
        wb.close()
        
        compiler = ModelCompiler()
        
        # Load without ignore_hidden (should include hidden sheet)
        model_with_hidden = compiler.read_and_parse_archive(excel_file)
        
        # Load with ignore_hidden=True
        model_no_hidden = compiler.read_and_parse_archive(
            excel_file,
            ignore_hidden=True
        )
        
        # Get sheet names
        def get_sheet_names(model):
            cell_addresses = list(model.cells.keys())
            return set(addr.split('!')[0] for addr in cell_addresses if '!' in addr)
        
        sheets_with_hidden = get_sheet_names(model_with_hidden)
        sheets_no_hidden = get_sheet_names(model_no_hidden)
        
        print(f"Sheets with hidden included: {sorted(sheets_with_hidden)}")
        print(f"Sheets with hidden ignored: {sorted(sheets_no_hidden)}")
        print(f"Hidden sheet excluded: {'✓' if 'Config' not in sheets_no_hidden else '✗'}")
        
    finally:
        os.unlink(excel_file)


if __name__ == '__main__':
    print("xlcalculator - Model Focusing Example")
    print("=" * 50)
    print("This example demonstrates focusing capabilities:")
    print("1. ignore_sheets parameter - exclude specific sheets")
    print("2. ignore_hidden parameter - exclude hidden sheets")
    print("=" * 50)
    
    example_full_model_analysis()
    example_focused_model_with_ignore_sheets()
    example_scenario_analysis_with_focusing()
    example_performance_comparison()
    example_ignore_hidden_sheets()
    
    print("\n✅ Example completed successfully!")
    print("\nKey Takeaways:")
    print("• ignore_sheets parameter excludes specific sheets during loading")
    print("• ignore_hidden parameter excludes hidden sheets during loading")
    print("• Focused models reduce memory usage and improve performance")
    print("• All calculations remain accurate in focused models")
    print("• Model focusing is ideal for large Excel files with unnecessary data")