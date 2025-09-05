#!/usr/bin/env python3
"""
Excel File Template Generator for Integration Tests

This script creates Excel files with formulas and test data for validating
xlcalculator functions against Excel behavior.
"""

import openpyxl
from openpyxl import Workbook
import os

def create_xlookup_excel():
    """Create XLOOKUP.xlsx with comprehensive test scenarios and calculated values."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Test data setup
    ws['A1'] = 'Fruit'
    ws['B1'] = 'Price'
    ws['A2'] = 'Apple'
    ws['B2'] = 10
    ws['A3'] = 'Banana'
    ws['B3'] = 20
    ws['A4'] = 'Cherry'
    ws['B4'] = 30
    ws['A5'] = 'Date'
    ws['B5'] = 40
    
    # Sorted numbers for approximate match testing
    ws['D1'] = 'Score'
    ws['E1'] = 'Grade'
    ws['D2'] = 10
    ws['E2'] = 'F'
    ws['D3'] = 20
    ws['E3'] = 'D'
    ws['D4'] = 30
    ws['E4'] = 'C'
    ws['D5'] = 40
    ws['E5'] = 'B'
    ws['D6'] = 50
    ws['E6'] = 'A'
    
    # Duplicate values for reverse search testing
    ws['G1'] = 'Item'
    ws['H1'] = 'Position'
    ws['G2'] = 'A'
    ws['H2'] = 1
    ws['G3'] = 'B'
    ws['H3'] = 2
    ws['G4'] = 'A'
    ws['H4'] = 3
    ws['G5'] = 'C'
    ws['H5'] = 4
    ws['G6'] = 'A'
    ws['H6'] = 5
    
    # Test formulas with expected calculated values
    # For integration tests, we store both formula and expected result
    ws['A8'] = '=XLOOKUP("Apple", A2:A5, B2:B5)'  # Should return 10
    ws['A8'].value = 10  # Store expected result
    
    ws['A9'] = '=XLOOKUP("Orange", A2:A5, B2:B5, "Not Found")'  # Should return "Not Found"
    ws['A9'].value = "Not Found"
    
    ws['A10'] = '=XLOOKUP("Cherry", A2:A5, B2:B5)'  # Should return 30
    ws['A10'].value = 30
    
    # Approximate match tests
    ws['A12'] = '=XLOOKUP(25, D2:D6, E2:E6, , -1)'  # Should return "D" (next smallest: 20)
    ws['A12'].value = "D"
    
    ws['A13'] = '=XLOOKUP(15, D2:D6, E2:E6, , 1)'   # Should return "D" (next largest: 20)
    ws['A13'].value = "D"
    
    ws['A14'] = '=XLOOKUP(30, D2:D6, E2:E6, , 0)'   # Should return "C" (exact match)
    ws['A14'].value = "C"
    
    # Wildcard match tests
    ws['A16'] = '=XLOOKUP("App*", A2:A5, B2:B5, , 2)'    # Should return 10
    ws['A16'].value = 10
    
    ws['A17'] = '=XLOOKUP("Ban?na", A2:A5, B2:B5, , 2)'  # Should return 20
    ws['A17'].value = 20
    
    ws['A18'] = '=XLOOKUP("*erry", A2:A5, B2:B5, , 2)'   # Should return 30
    ws['A18'].value = 30
    
    # Reverse search tests
    ws['A20'] = '=XLOOKUP("A", G2:G6, H2:H6, , 0, 1)'   # Should return 1 (first)
    ws['A20'].value = 1
    
    ws['A21'] = '=XLOOKUP("A", G2:G6, H2:H6, , 0, -1)'  # Should return 5 (last)
    ws['A21'].value = 5
    
    # Binary search tests (sorted data)
    ws['A23'] = '=XLOOKUP(30, D2:D6, E2:E6, , 0, 2)'    # Should return "C"
    ws['A23'].value = "C"
    
    ws['A24'] = '=XLOOKUP(20, D2:D6, E2:E6, , 0, 2)'    # Should return "D"
    ws['A24'].value = "D"
    
    # Error cases - store the error value
    ws['A26'] = '=XLOOKUP("Grape", A2:A5, B2:B5)'       # Should return #N/A
    # Note: We'll handle error values in the test
    
    # Horizontal array test
    ws['A28'] = 'Apple'
    ws['B28'] = 'Banana'
    ws['C28'] = 'Cherry'
    ws['A29'] = 100
    ws['B29'] = 200
    ws['C29'] = 300
    ws['A30'] = '=XLOOKUP("Banana", A28:C28, A29:C29)'  # Should return 200
    ws['A30'].value = 200
    
    return wb

def create_logical_excel():
    """Create LOGICAL.xlsx with AND, OR, TRUE, FALSE tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Test data
    ws['A1'] = True
    ws['B1'] = False
    ws['C1'] = 5
    ws['D1'] = 10
    ws['E1'] = 0
    
    # AND function tests
    ws['A3'] = '=AND(TRUE, TRUE)'
    ws['A4'] = '=AND(TRUE, FALSE)'
    ws['A5'] = '=AND(FALSE, FALSE)'
    ws['A6'] = '=AND(A1, B1)'  # Reference to cells
    ws['A7'] = '=AND(C1>0, D1>5)'  # Logical expressions
    ws['A8'] = '=AND(C1>0, D1>5, E1=0)'  # Multiple conditions
    
    # OR function tests
    ws['B3'] = '=OR(TRUE, TRUE)'
    ws['B4'] = '=OR(TRUE, FALSE)'
    ws['B5'] = '=OR(FALSE, FALSE)'
    ws['B6'] = '=OR(A1, B1)'  # Reference to cells
    ws['B7'] = '=OR(C1>10, D1>5)'  # Logical expressions
    ws['B8'] = '=OR(C1>10, D1>15, E1>0)'  # Multiple conditions
    
    # TRUE and FALSE constants
    ws['C3'] = '=TRUE()'
    ws['C4'] = '=FALSE()'
    
    # Nested logical functions
    ws['D3'] = '=AND(OR(A1, B1), NOT(E1>0))'
    ws['D4'] = '=OR(AND(A1, B1), AND(C1>0, D1>0))'
    
    # Edge cases
    ws['E3'] = '=AND()'  # Empty AND (should be TRUE)
    ws['E4'] = '=OR()'   # Empty OR (should be FALSE)
    
    return wb

def create_information_excel():
    """Create INFORMATION.xlsx with IS* function tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Test data with various types
    ws['A1'] = 123        # Number
    ws['A2'] = "Hello"    # Text
    ws['A3'] = True       # Boolean
    ws['A4'] = None       # Empty/Blank
    ws['A5'] = '=1/0'     # Error formula
    ws['A6'] = '=NA()'    # #N/A error
    ws['A7'] = 4          # Even number
    ws['A8'] = 5          # Odd number
    ws['A9'] = 0          # Zero
    ws['A10'] = -3        # Negative odd
    
    # ISNUMBER tests
    ws['B1'] = '=ISNUMBER(A1)'   # Number
    ws['B2'] = '=ISNUMBER(A2)'   # Text
    ws['B3'] = '=ISNUMBER(A3)'   # Boolean
    ws['B4'] = '=ISNUMBER(A4)'   # Blank
    ws['B5'] = '=ISNUMBER(123)'  # Direct number
    
    # ISTEXT tests
    ws['C1'] = '=ISTEXT(A1)'     # Number
    ws['C2'] = '=ISTEXT(A2)'     # Text
    ws['C3'] = '=ISTEXT(A3)'     # Boolean
    ws['C4'] = '=ISTEXT(A4)'     # Blank
    ws['C5'] = '=ISTEXT("test")' # Direct text
    
    # ISBLANK tests
    ws['D1'] = '=ISBLANK(A1)'    # Number
    ws['D2'] = '=ISBLANK(A2)'    # Text
    ws['D4'] = '=ISBLANK(A4)'    # Blank
    ws['D5'] = '=ISBLANK("")'    # Empty string
    
    # ISERROR tests
    ws['E1'] = '=ISERROR(A1)'    # Number
    ws['E5'] = '=ISERROR(A5)'    # Error
    ws['E6'] = '=ISERROR(A6)'    # #N/A
    ws['E7'] = '=ISERROR(1/0)'   # Direct error
    
    # ISNA tests
    ws['F6'] = '=ISNA(A6)'       # #N/A error
    ws['F5'] = '=ISNA(A5)'       # Other error
    ws['F1'] = '=ISNA(A1)'       # Number
    
    # ISERR tests (errors except #N/A)
    ws['G5'] = '=ISERR(A5)'      # #DIV/0! error
    ws['G6'] = '=ISERR(A6)'      # #N/A error
    ws['G1'] = '=ISERR(A1)'      # Number
    
    # ISEVEN tests
    ws['H7'] = '=ISEVEN(A7)'     # Even number
    ws['H8'] = '=ISEVEN(A8)'     # Odd number
    ws['H9'] = '=ISEVEN(A9)'     # Zero
    ws['H10'] = '=ISEVEN(A10)'   # Negative odd
    
    # ISODD tests
    ws['I7'] = '=ISODD(A7)'      # Even number
    ws['I8'] = '=ISODD(A8)'      # Odd number
    ws['I9'] = '=ISODD(A9)'      # Zero
    ws['I10'] = '=ISODD(A10)'    # Negative odd
    
    # NA function test
    ws['J1'] = '=NA()'           # Generate #N/A error
    
    return wb

def create_math_excel():
    """Create MATH.xlsx with mathematical function tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Test data
    ws['A1'] = 3.7
    ws['A2'] = -2.3
    ws['A3'] = 0
    ws['A4'] = 100
    ws['A5'] = 2.71828  # Approximately e
    
    # FLOOR tests
    ws['B1'] = '=FLOOR(A1, 1)'     # 3.7 -> 3
    ws['B2'] = '=FLOOR(A2, 1)'     # -2.3 -> -3
    ws['B3'] = '=FLOOR(A1, 0.5)'   # 3.7 -> 3.5
    ws['B4'] = '=FLOOR(A4, 10)'    # 100 -> 100
    
    # TRUNC tests
    ws['C1'] = '=TRUNC(A1)'        # 3.7 -> 3
    ws['C2'] = '=TRUNC(A2)'        # -2.3 -> -2
    ws['C3'] = '=TRUNC(A1, 1)'     # 3.7 -> 3.7
    ws['C4'] = '=TRUNC(A4, -1)'    # 100 -> 100
    
    # SIGN tests
    ws['D1'] = '=SIGN(A1)'         # Positive -> 1
    ws['D2'] = '=SIGN(A2)'         # Negative -> -1
    ws['D3'] = '=SIGN(A3)'         # Zero -> 0
    
    # LOG tests
    ws['E1'] = '=LOG(A4)'          # LOG base 10
    ws['E2'] = '=LOG(A4, 2)'       # LOG base 2
    ws['E3'] = '=LOG(A5, EXP(1))'  # Natural log
    
    # LOG10 tests
    ws['F1'] = '=LOG10(A4)'        # LOG10(100) = 2
    ws['F2'] = '=LOG10(1000)'      # LOG10(1000) = 3
    
    # EXP tests
    ws['G1'] = '=EXP(0)'           # e^0 = 1
    ws['G2'] = '=EXP(1)'           # e^1 = e
    ws['G3'] = '=EXP(2)'           # e^2
    
    return wb

def create_text_excel():
    """Create TEXT.xlsx with text function tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Test data
    ws['A1'] = "Hello World"
    ws['A2'] = "  Excel Functions  "
    ws['A3'] = "UPPERCASE"
    ws['A4'] = "lowercase"
    ws['A5'] = "MiXeD cAsE"
    ws['A6'] = "Replace This Text"
    
    # LEFT tests
    ws['B1'] = '=LEFT(A1, 5)'      # "Hello"
    ws['B2'] = '=LEFT(A1, 1)'      # "H"
    ws['B3'] = '=LEFT(A1)'         # Default 1 char
    
    # UPPER tests
    ws['C3'] = '=UPPER(A3)'        # Already uppercase
    ws['C4'] = '=UPPER(A4)'        # Convert lowercase
    ws['C5'] = '=UPPER(A5)'        # Convert mixed case
    
    # LOWER tests
    ws['D3'] = '=LOWER(A3)'        # Convert uppercase
    ws['D4'] = '=LOWER(A4)'        # Already lowercase
    ws['D5'] = '=LOWER(A5)'        # Convert mixed case
    
    # TRIM tests
    ws['E2'] = '=TRIM(A2)'         # Remove leading/trailing spaces
    ws['E1'] = '=TRIM(A1)'         # No extra spaces
    
    # REPLACE tests
    ws['F6'] = '=REPLACE(A6, 9, 4, "That")'  # Replace "This" with "That"
    ws['F1'] = '=REPLACE(A1, 1, 5, "Hi")'    # Replace "Hello" with "Hi"
    
    return wb

def save_excel_files():
    """Save all Excel files to the resources directory."""
    resources_dir = "tests/resources"
    
    # Ensure resources directory exists
    os.makedirs(resources_dir, exist_ok=True)
    
    files_to_create = [
        ("XLOOKUP.xlsx", create_xlookup_excel),
        ("LOGICAL.xlsx", create_logical_excel),
        ("INFORMATION.xlsx", create_information_excel),
        ("MATH.xlsx", create_math_excel),
        ("TEXT.xlsx", create_text_excel),
    ]
    
    created_files = []
    for filename, create_func in files_to_create:
        filepath = os.path.join(resources_dir, filename)
        try:
            wb = create_func()
            wb.save(filepath)
            created_files.append(filename)
            print(f"✅ Created {filename}")
        except Exception as e:
            print(f"❌ Failed to create {filename}: {e}")
    
    return created_files

if __name__ == "__main__":
    print("Creating Excel files for integration tests...")
    created = save_excel_files()
    print(f"\nSuccessfully created {len(created)} Excel files:")
    for file in created:
        print(f"  - {file}")