#!/usr/bin/env python3
"""
Create a proper Excel file for testing sheet context behavior.

This creates a minimal Excel file with specific formulas to test
how implicit references should resolve to the correct sheet context.
"""

import openpyxl
import os

def create_sheet_context_test_file():
    """Create Excel file with multi-sheet formulas for context testing."""
    
    # Create workbook with multiple sheets
    wb = openpyxl.Workbook()
    
    # Remove default sheet and create named sheets
    wb.remove(wb.active)
    sheet1 = wb.create_sheet("Sheet1")
    sheet2 = wb.create_sheet("Sheet2")
    
    # Sheet1 data and formulas
    sheet1['A1'] = 10
    sheet1['A2'] = 20
    sheet1['A3'] = 30
    sheet1['B1'] = 5
    sheet1['B2'] = 15
    sheet1['B3'] = 25
    
    # Sheet1 formula with implicit reference (should resolve to Sheet1!A1:A3)
    sheet1['C1'] = '=SUM(A1:A3)'  # Should sum Sheet1!A1:A3 = 60
    
    # Sheet1 formula with mixed references
    sheet1['C2'] = '=A1 + Sheet2!A1'  # Sheet1!A1 + Sheet2!A1
    
    # Sheet2 data and formulas  
    sheet2['A1'] = 100
    sheet2['A2'] = 200
    sheet2['A3'] = 300
    sheet2['B1'] = 50
    sheet2['B2'] = 150
    sheet2['B3'] = 250
    
    # Sheet2 formula with implicit reference (should resolve to Sheet2!A1:A3)
    sheet2['C1'] = '=SUM(A1:A3)'  # Should sum Sheet2!A1:A3 = 600
    
    # Sheet2 formula with mixed references
    sheet2['C2'] = '=A1 + Sheet1!A1'  # Sheet2!A1 + Sheet1!A1
    
    # Cross-sheet reference from Sheet1 to Sheet2
    sheet1['D1'] = '=Sheet2!C1'  # Reference to Sheet2's SUM formula
    
    # Cross-sheet reference from Sheet2 to Sheet1
    sheet2['D1'] = '=Sheet1!C1'  # Reference to Sheet1's SUM formula
    
    # Save the file
    resource_dir = os.path.join(os.path.dirname(__file__), 'resources')
    filename = os.path.join(resource_dir, 'sheet_context_test.xlsx')
    wb.save(filename)
    
    print(f"Created test file: {filename}")
    print("\nExpected behavior:")
    print("Sheet1!C1 = SUM(A1:A3) should resolve to SUM(Sheet1!A1:A3) = 60")
    print("Sheet2!C1 = SUM(A1:A3) should resolve to SUM(Sheet2!A1:A3) = 600")
    print("Sheet1!C2 = A1 + Sheet2!A1 should resolve to Sheet1!A1 + Sheet2!A1 = 110")
    print("Sheet2!C2 = A1 + Sheet1!A1 should resolve to Sheet2!A1 + Sheet1!A1 = 110")
    
    return filename

if __name__ == '__main__':
    create_sheet_context_test_file()