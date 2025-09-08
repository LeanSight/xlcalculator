#!/usr/bin/env python3
"""
Generate Excel files from JSON test configuration using openpyxl.
Creates Excel workbooks with test data and formulas for dynamic ranges testing.
"""

import argparse
import json
from pathlib import Path
from typing import Dict, Any, List
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def load_json_config(json_path: str) -> Dict[str, Any]:
    """Load and parse JSON test configuration."""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def create_data_sheet(wb: Workbook, data_config: Dict[str, Any]) -> None:
    """Create the Data sheet with test data."""
    # Create or get Data sheet
    if 'Data' in wb.sheetnames:
        ws = wb['Data']
    else:
        ws = wb.create_sheet('Data')
    
    # Clear existing data
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    
    # Add headers
    headers = data_config['headers']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Add data rows
    for row_idx, row_data in enumerate(data_config['rows'], 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    print(f"Created Data sheet with {len(headers)} columns and {len(data_config['rows'])} data rows")


def create_tests_sheet(wb: Workbook, levels: List[Dict], auxiliary_data: Dict[str, Any]) -> None:
    """Create the Tests sheet with formulas and auxiliary data."""
    # Create or get Tests sheet
    if 'Tests' in wb.sheetnames:
        ws = wb['Tests']
    else:
        ws = wb.create_sheet('Tests')
    
    # Clear existing data
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    
    # Add auxiliary data first
    for cell_ref, value in auxiliary_data.items():
        # Parse cell reference (e.g., "P1", "Z1")
        col_letter = ''.join(c for c in cell_ref if c.isalpha())
        row_num = int(''.join(c for c in cell_ref if c.isdigit()))
        
        # Convert column letter to number
        col_num = openpyxl.utils.column_index_from_string(col_letter)
        
        ws.cell(row=row_num, column=col_num, value=value)
        print(f"Added auxiliary data: {cell_ref} = {value}")
    
    # Add test formulas
    formula_count = 0
    for level in levels:
        for test_case in level['test_cases']:
            cell_ref = test_case['cell']
            formula = test_case['formula']
            
            # Parse cell reference
            col_letter = ''.join(c for c in cell_ref if c.isalpha())
            row_num = int(''.join(c for c in cell_ref if c.isdigit()))
            
            # Convert column letter to number
            col_num = openpyxl.utils.column_index_from_string(col_letter)
            
            # Set formula (remove = if present)
            formula_clean = formula[1:] if formula.startswith('=') else formula
            ws.cell(row=row_num, column=col_num, value=f"={formula_clean}")
            formula_count += 1
            
            print(f"Added formula: {cell_ref} = {formula}")
    
    print(f"Created Tests sheet with {len(auxiliary_data)} auxiliary data items and {formula_count} formulas")


def create_excel_file(config: Dict[str, Any], output_path: str) -> None:
    """Create Excel file from JSON configuration."""
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Data sheet
    data_config = config['data_sheet']
    create_data_sheet(wb, data_config)
    
    # Create Tests sheet
    levels = config['levels']
    auxiliary_data = config['auxiliary_data']
    create_tests_sheet(wb, levels, auxiliary_data)
    
    # Set Data as active sheet
    wb.active = wb['Data']
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel file saved: {output_path}")


def main(json_path: str, output_dir: str) -> None:
    """Main function to generate Excel from JSON."""
    # Load configuration
    config = load_json_config(json_path)
    
    # Get output filename
    excel_filename = config['generation_config'].get('excel_filename', 'dynamic_ranges.xlsx')
    output_path = Path(output_dir) / excel_filename
    
    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Create Excel file
    create_excel_file(config, str(output_path))
    
    print(f"Successfully generated Excel file: {output_path}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Generate Excel files from JSON test configuration using openpyxl')
    parser.add_argument('json_path', help='Path to JSON test configuration file')
    parser.add_argument('output_dir', help='Output directory for Excel file')
    
    args = parser.parse_args()
    main(args.json_path, args.output_dir)