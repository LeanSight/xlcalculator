#!/usr/bin/env python3
"""
Generate Excel test files from JSON test configuration by category.
Creates separate Excel workbooks for each category with Data and Tests sheets.
Auto-detects platform: xlwings on Windows, openpyxl on Unix.
"""

import argparse
import platform
from pathlib import Path
from typing import Dict, List, Any, Set
from collections import defaultdict

try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from json_to_tests_utils import (
    load_json_config, extract_test_levels, extract_data_config,
    extract_auxiliary_data, extract_metadata, extract_generation_config,
    validate_json_and_output_dir, count_total_test_cases, TestLevel
)


def detect_excel_method() -> str:
    """Auto-detect best Excel generation method based on platform."""
    is_windows = platform.system() == "Windows"
    
    if is_windows and XLWINGS_AVAILABLE:
        try:
            app = xw.App(visible=False)
            app.quit()
            return "xlwings"
        except Exception:
            pass
    
    if OPENPYXL_AVAILABLE:
        return "openpyxl"
    
    raise RuntimeError("No Excel generation method available. Install xlwings (Windows) or openpyxl (Unix)")


def group_levels_by_category(levels: List[TestLevel]) -> Dict[str, List[TestLevel]]:
    """Group test levels by category."""
    categories = defaultdict(list)
    for level in levels:
        category = level.category
        categories[category].append(level)
    return dict(categories)


def get_auxiliary_data_for_category(aux_data: Dict[str, Any], levels: List[TestLevel]) -> Dict[str, Any]:
    """Extract only auxiliary data needed for specific category."""
    needed_refs = set()
    
    # Scan formulas for auxiliary data references
    for level in levels:
        for case in level.test_cases:
            formula = case.formula
            # Look for P1, P2, etc. references
            for key in aux_data.keys():
                if key in formula:
                    needed_refs.add(key)
    
    return {key: value for key, value in aux_data.items() if key in needed_refs}


def get_category_filename(category: str, metadata: Dict[str, Any]) -> str:
    """Generate filename for category."""
    base_name = metadata.get("title", "test_suite").replace(" ", "_").replace("-", "_").lower()
    return f"{category}.xlsx"


def create_data_sheet(wb, data_config: Dict[str, Any]) -> None:
    """Create and populate Data sheet with test data."""
    data_sheet = wb.sheets[0]
    data_sheet.name = "Data"
    
    # Add headers
    headers = data_config["headers"]
    for i, header in enumerate(headers, 1):
        data_sheet.cells(1, i).value = header
    
    # Add data rows
    for row_idx, row_data in enumerate(data_config["rows"], 2):
        for col_idx, value in enumerate(row_data, 1):
            data_sheet.cells(row_idx, col_idx).value = value


def create_auxiliary_data(tests_sheet, aux_data: Dict[str, Any]) -> None:
    """Add auxiliary data for INDIRECT tests."""
    for cell, value in aux_data.items():
        tests_sheet[cell].value = value


def add_formula_to_sheet(tests_sheet, cell: str, formula: str) -> None:
    """Add formula to specific cell with error handling."""
    try:
        tests_sheet[cell].formula = formula
    except Exception as e:
        print(f"Warning: Failed to add formula {cell}: {formula} - {e}")


def populate_test_formulas(tests_sheet, levels: List[TestLevel]) -> int:
    """Add all test formulas to Tests sheet."""
    formula_count = 0
    
    for level in levels:
        for case in level.test_cases:
            add_formula_to_sheet(tests_sheet, case.cell, case.formula)
            formula_count += 1
    
    return formula_count


def add_level_labels(tests_sheet, levels: List[TestLevel], gen_config: Dict[str, Any]) -> None:
    """Add descriptive labels for each test level."""
    label_row = gen_config.get("label_row", 20)
    
    for level in levels:
        # Extract column from first test case
        if level.test_cases:
            first_cell = level.test_cases[0].cell
            column = first_cell.rstrip('0123456789')
            label_cell = f"{column}{label_row}"
            tests_sheet[label_cell].value = f"{level.level}: {level.title}"


def create_tests_sheet(wb, levels: List[TestLevel], aux_data: Dict[str, Any], gen_config: Dict[str, Any]) -> int:
    """Create and populate Tests sheet with formulas."""
    tests_sheet = wb.sheets.add("Tests")
    
    # Add auxiliary data first
    create_auxiliary_data(tests_sheet, aux_data)
    
    # Add test formulas
    formula_count = populate_test_formulas(tests_sheet, levels)
    
    # Add descriptive labels
    add_level_labels(tests_sheet, levels, gen_config)
    
    return formula_count


def force_calculation(wb) -> None:
    """Force Excel to calculate all formulas."""
    try:
        wb.app.calculate()
        print("Excel calculation completed")
    except Exception as e:
        print(f"Calculation warning: {e}")


def create_excel_workbook_xlwings(category: str, levels: List[TestLevel], data_config: Dict[str, Any], 
                                  aux_data: Dict[str, Any], gen_config: Dict[str, Any], output_path: Path) -> None:
    """Create Excel workbook using xlwings."""
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    
    try:
        wb = app.books.add()
        
        # Create sheets
        create_data_sheet(wb, data_config)
        formula_count = create_tests_sheet(wb, levels, aux_data, gen_config)
        
        # Calculate and save
        force_calculation(wb)
        wb.save(str(output_path))
        
        print(f"Excel created (xlwings): {output_path}")
        print(f"Category: {category}, {formula_count} formulas across {len(levels)} levels")
        
    except Exception as e:
        print(f"Excel creation failed: {e}")
        raise
    finally:
        try:
            if 'wb' in locals():
                wb.close()
        except:
            pass
        try:
            app.quit()
        except:
            pass


def create_excel_workbook_openpyxl(category: str, levels: List[TestLevel], data_config: Dict[str, Any], 
                                   aux_data: Dict[str, Any], gen_config: Dict[str, Any], output_path: Path) -> None:
    """Create Excel workbook using openpyxl."""
    wb = Workbook()
    
    # Create Data sheet
    data_sheet = wb.active
    data_sheet.title = "Data"
    
    # Add headers
    headers = data_config["headers"]
    for i, header in enumerate(headers, 1):
        data_sheet.cell(row=1, column=i, value=header)
    
    # Add data rows
    for row_idx, row_data in enumerate(data_config["rows"], 2):
        for col_idx, value in enumerate(row_data, 1):
            data_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Tests sheet
    tests_sheet = wb.create_sheet("Tests")
    
    # Add auxiliary data
    for cell, value in aux_data.items():
        tests_sheet[cell] = value
    
    # Add test formulas
    formula_count = 0
    for level in levels:
        for case in level.test_cases:
            tests_sheet[case.cell] = case.formula
            formula_count += 1
    
    # Add level labels
    label_row = gen_config.get("label_row", 20)
    for level in levels:
        if level.test_cases:
            first_cell = level.test_cases[0].cell
            column = first_cell.rstrip('0123456789')
            label_cell = f"{column}{label_row}"
            tests_sheet[label_cell] = f"{level.level}: {level.title}"
    
    # Save workbook
    wb.save(str(output_path))
    
    print(f"Excel created (openpyxl): {output_path}")
    print(f"Category: {category}, {formula_count} formulas across {len(levels)} levels")
    print("Note: Formulas not calculated - use Excel to open and calculate")


def generate_summary_report(config: Dict[str, Any]) -> None:
    """Print summary of Excel generation."""
    metadata = extract_metadata(config)
    gen_config = extract_generation_config(config)
    levels = extract_test_levels(config)
    total_cases = count_total_test_cases(levels)
    
    title = metadata.get('title', gen_config.get('class_name', 'Test Suite'))
    
    print("\n" + "="*60)
    print("EXCEL GENERATION SUMMARY")
    print("="*60)
    print(f"Title: {title}")
    print(f"Total test cases: {total_cases}")
    print(f"Test levels: {len(levels)}")
    print(f"Source: {metadata.get('source', 'JSON configuration')}")
    
    print("\nTEST CATEGORIES:")
    for level in levels:
        case_count = len(level.test_cases)
        print(f"   - {level.title}: {case_count} cases")


def generate_excel_by_category(config: Dict[str, Any], output_path: Path, method: str) -> None:
    """Generate Excel files by category."""
    levels = extract_test_levels(config)
    data_config = extract_data_config(config)
    gen_config = extract_generation_config(config)
    aux_data = extract_auxiliary_data(config)
    metadata = extract_metadata(config)
    
    # Group levels by category
    categories = group_levels_by_category(levels)
    
    print(f"Generating {len(categories)} Excel files by category using {method}...")
    
    for category, category_levels in categories.items():
        # Get auxiliary data needed for this category
        category_aux_data = get_auxiliary_data_for_category(aux_data, category_levels)
        
        # Generate filename
        excel_filename = get_category_filename(category, metadata)
        excel_file_path = output_path / excel_filename
        
        # Create Excel file
        if method == "xlwings":
            create_excel_workbook_xlwings(category, category_levels, data_config, 
                                        category_aux_data, gen_config, excel_file_path)
        else:
            create_excel_workbook_openpyxl(category, category_levels, data_config, 
                                         category_aux_data, gen_config, excel_file_path)


def main(json_path: str, output_dir: str, method: str = None) -> None:
    """Generate Excel files by category from JSON configuration."""
    json_file, output_path = validate_json_and_output_dir(json_path, output_dir)
    
    # Auto-detect method if not specified
    if method is None:
        method = detect_excel_method()
    
    print(f"Using {method} for Excel generation")
    
    config = load_json_config(str(json_file))
    
    generate_excel_by_category(config, output_path, method)
    generate_summary_report(config)
    
    print("\nNEXT STEPS:")
    print("1. Use generated Excel files for integration testing")
    print("2. Run tests to verify Excel compatibility")
    print("3. Implement functions using red-green-refactor strategy")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate Excel test files by category from JSON configuration"
    )
    parser.add_argument("json_path", help="Path to JSON test configuration file")
    parser.add_argument("output_dir", help="Output directory for generated Excel files")
    parser.add_argument("--method", choices=["xlwings", "openpyxl"], 
                       help="Excel generation method (auto-detected if not specified)")
    
    args = parser.parse_args()
    main(args.json_path, args.output_dir, args.method)